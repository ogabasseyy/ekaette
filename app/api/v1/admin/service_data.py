"""Data import/export/purge service functions.

Extracted from main.py as Phase B4 of modularization. Zero behavior changes.
"""

from __future__ import annotations

import asyncio
import csv
from datetime import datetime, timedelta, timezone
import io
import json
import logging
import re
from urllib.parse import parse_qs, urlparse

import httpx

logger = logging.getLogger(__name__)

from app.api.v1.admin.runtime import runtime as _m

from app.api.v1.admin.firestore_helpers import (
    _batch_delete_documents,
    _batch_set_documents,
    _doc_delete,
    _doc_get,
)
from app.api.v1.admin.service_companies import _admin_company_response
from app.api.v1.admin.service_knowledge import (
    _collect_query_docs,
    _list_company_collection_docs,
)

_TRUTHY = {"1", "true", "yes", "y", "on", "available", "in_stock", "instock"}
_FALSY = {"0", "false", "no", "n", "off", "out_of_stock", "outofstock", "sold_out"}

_INVENTORY_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "id": ("id", "sku", "product_id", "productid", "item_id", "itemid"),
    "name": ("name", "product_name", "productname", "item_name", "itemname", "title"),
    "category": ("category", "product_category", "productcategory", "department"),
    "price": ("price", "unit_price", "unitprice", "amount"),
    "currency": ("currency", "currency_code", "currencycode"),
    "in_stock": ("in_stock", "instock", "available", "availability", "stock_status", "stockstatus"),
    "quantity": ("quantity", "qty", "stock", "inventory", "on_hand", "onhand"),
    "description": ("description", "details", "summary"),
    "brand": ("brand", "maker"),
    "warehouse": ("warehouse", "location", "warehouse_location", "warehouselocation"),
}

_DEMO_SEED_VERSION = "electronics-v1"

_DEMO_SEED_KNOWLEDGE: list[dict[str, object]] = [
    {
        "title": "Support hours",
        "text": (
            "Customer support is available daily from 9 AM to 7 PM (WAT). "
            "Online chat is available 24/7."
        ),
        "tags": ["support", "hours"],
        "source": "seed",
    },
    {
        "title": "Pickup policy",
        "text": (
            "Same-day pickup is available for confirmed bookings made before 2 PM. "
            "Next-day pickup is guaranteed for all other bookings. Riders deliver within "
            "Lagos and Abuja metro areas only."
        ),
        "tags": ["pickup", "policy", "delivery"],
        "source": "seed",
    },
    {
        "title": "Trade-in warranty",
        "text": (
            "All trade-in devices come with a 30-day quality guarantee. If the device "
            "condition is misgraded, customers can request a free re-evaluation within "
            "14 days of trade-in."
        ),
        "tags": ["warranty", "policy", "trade-in"],
        "source": "seed",
    },
    {
        "title": "Payment methods",
        "text": (
            "We accept bank transfers, debit cards, USSD, and cash on pickup. Mobile money "
            "(OPay, PalmPay) is accepted for amounts under NGN 500,000. International cards "
            "are supported for online orders."
        ),
        "tags": ["payment", "policy"],
        "source": "seed",
    },
]

_DEMO_SEED_CONNECTOR: dict[str, object] = {
    "connectorId": "crm",
    "provider": "mock",
    "enabled": True,
    "capabilities": ["read"],
    "config": {},
}

_DEMO_SEED_PRODUCTS: list[dict[str, object]] = [
    {"id": "prod-iphone-15-pro-max", "name": "iPhone 15 Pro Max", "price": 950000, "currency": "NGN", "category": "smartphones", "brand": "Apple", "in_stock": True},
    {"id": "prod-iphone-15-pro", "name": "iPhone 15 Pro", "price": 850000, "currency": "NGN", "category": "smartphones", "brand": "Apple", "in_stock": True},
    {"id": "prod-iphone-14", "name": "iPhone 14", "price": 520000, "currency": "NGN", "category": "smartphones", "brand": "Apple", "in_stock": True},
    {"id": "prod-samsung-s24-ultra", "name": "Samsung Galaxy S24 Ultra", "price": 780000, "currency": "NGN", "category": "smartphones", "brand": "Samsung", "in_stock": True},
    {"id": "prod-samsung-s24", "name": "Samsung Galaxy S24", "price": 620000, "currency": "NGN", "category": "smartphones", "brand": "Samsung", "in_stock": True},
    {"id": "prod-google-pixel-8", "name": "Google Pixel 8", "price": 450000, "currency": "NGN", "category": "smartphones", "brand": "Google", "in_stock": False},
    {"id": "prod-ipad-air-m2", "name": "iPad Air M2", "price": 680000, "currency": "NGN", "category": "tablets", "brand": "Apple", "in_stock": True},
    {"id": "prod-samsung-tab-s9", "name": "Samsung Galaxy Tab S9", "price": 520000, "currency": "NGN", "category": "tablets", "brand": "Samsung", "in_stock": True},
    {"id": "prod-macbook-air-m3", "name": "MacBook Air M3", "price": 1250000, "currency": "NGN", "category": "laptops", "brand": "Apple", "in_stock": True},
    {"id": "prod-airpods-pro-2", "name": "AirPods Pro 2", "price": 180000, "currency": "NGN", "category": "accessories", "brand": "Apple", "in_stock": True},
    {"id": "prod-cctv-bundle-4ch", "name": "CCTV Security Bundle (4-Camera + DVR)", "price": 980000, "currency": "NGN", "category": "security", "brand": "Hikvision", "in_stock": True},
    {"id": "prod-anker-powerbank-26k", "name": "Anker PowerCore 26800", "price": 32000, "currency": "NGN", "category": "accessories", "brand": "Anker", "in_stock": True},
]

_DEMO_SEED_BOOKING_SLOTS: list[dict[str, object]] = [
    {"id": "slot-elec-0301-10-ikeja", "date": "2026-03-01", "time": "10:00", "location": "Lagos - Ikeja", "slot_type": "pickup", "available": True},
    {"id": "slot-elec-0301-14-ikeja", "date": "2026-03-01", "time": "14:00", "location": "Lagos - Ikeja", "slot_type": "pickup", "available": True},
    {"id": "slot-elec-0308-10-ikeja", "date": "2026-03-08", "time": "10:00", "location": "Lagos - Ikeja", "slot_type": "pickup", "available": True},
    {"id": "slot-elec-0308-14-wuse", "date": "2026-03-08", "time": "14:00", "location": "Abuja - Wuse", "slot_type": "pickup", "available": True},
    {"id": "slot-elec-0315-10-ikeja", "date": "2026-03-15", "time": "10:00", "location": "Lagos - Ikeja", "slot_type": "drop-off", "available": True},
    {"id": "slot-elec-0315-16-wuse", "date": "2026-03-15", "time": "16:00", "location": "Abuja - Wuse", "slot_type": "drop-off", "available": True},
    {"id": "slot-elec-0322-10-ikeja", "date": "2026-03-22", "time": "10:00", "location": "Lagos - Ikeja", "slot_type": "pickup", "available": False},
    {"id": "slot-elec-0322-14-vi", "date": "2026-03-22", "time": "14:00", "location": "Lagos - Victoria Island", "slot_type": "pickup", "available": True},
]


def _header_key(raw_value: object) -> str:
    text = str(raw_value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _slugify(raw_value: object) -> str:
    text = str(raw_value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def _json_response_error_message(response: object, *, fallback: str) -> str:
    body = getattr(response, "body", b"")
    if isinstance(body, bytes):
        try:
            parsed = json.loads(body.decode("utf-8"))
            if isinstance(parsed, dict) and isinstance(parsed.get("error"), str) and parsed["error"].strip():
                return parsed["error"].strip()
        except (json.JSONDecodeError, UnicodeDecodeError) as exc:
            # Response body may be non-JSON (for example, HTML proxy errors).
            logger.debug("Could not parse response body as JSON error payload: %s", exc)
    return fallback


def _demo_seed_knowledge_entries(*, data_tier: str) -> list[dict[str, object]]:
    now_iso = datetime.now(timezone.utc).isoformat()
    entries: list[dict[str, object]] = []
    for idx, item in enumerate(_DEMO_SEED_KNOWLEDGE, start=1):
        title = str(item.get("title") or f"Demo Knowledge {idx}").strip()
        slug = _slugify(title) or f"demo-{idx}"
        tags = item.get("tags")
        normalized_tags = (
            [str(tag).strip().lower() for tag in tags if str(tag).strip()]
            if isinstance(tags, list)
            else ["general"]
        )
        entries.append(
            {
                "id": f"kb-demo-{slug}"[:64],
                "title": title,
                "text": str(item.get("text") or "").strip(),
                "tags": normalized_tags or ["general"],
                "source": str(item.get("source") or "seed").strip().lower() or "seed",
                "data_tier": data_tier,
                "updated_at": now_iso,
            }
        )
    return entries


def _first_non_empty(row: dict[str, object], aliases: tuple[str, ...]) -> object | None:
    for key in aliases:
        value = row.get(key)
        if isinstance(value, str):
            if value.strip():
                return value.strip()
            continue
        if value is not None:
            return value
    return None


def _parse_float(raw_value: object) -> float | None:
    if isinstance(raw_value, (int, float)):
        return float(raw_value)
    if not isinstance(raw_value, str):
        return None
    candidate = raw_value.strip()
    if not candidate:
        return None
    candidate = candidate.replace(",", "")
    try:
        return float(candidate)
    except ValueError:
        return None


def _parse_int(raw_value: object) -> int | None:
    parsed = _parse_float(raw_value)
    if parsed is None:
        return None
    return int(parsed)


def _parse_bool(raw_value: object) -> bool | None:
    if isinstance(raw_value, bool):
        return raw_value
    if isinstance(raw_value, (int, float)):
        return bool(raw_value)
    if not isinstance(raw_value, str):
        return None
    normalized = _header_key(raw_value)
    if normalized in _TRUTHY:
        return True
    if normalized in _FALSY:
        return False
    return None


def _normalize_inventory_row(
    row: dict[str, object],
    *,
    index: int,
    data_tier: str,
) -> tuple[dict[str, object] | None, str | None]:
    normalized_row = {_header_key(key): value for key, value in row.items() if _header_key(key)}
    name = _first_non_empty(normalized_row, _INVENTORY_HEADER_ALIASES["name"])
    if not isinstance(name, str) or not name.strip():
        return None, f"row {index}: missing product name"
    name_value = name.strip()

    product_id_raw = _first_non_empty(normalized_row, _INVENTORY_HEADER_ALIASES["id"])
    product_id = str(product_id_raw).strip().lower() if product_id_raw is not None else ""
    if not product_id:
        slug = _slugify(name_value) or f"product-{index}"
        product_id = f"{slug}-{index}"

    price_raw = _first_non_empty(normalized_row, _INVENTORY_HEADER_ALIASES["price"])
    price = _parse_float(price_raw)
    if price is None or price < 0:
        return None, f"row {index}: missing or invalid price"

    quantity_raw = _first_non_empty(normalized_row, _INVENTORY_HEADER_ALIASES["quantity"])
    quantity = _parse_int(quantity_raw)

    in_stock_raw = _first_non_empty(normalized_row, _INVENTORY_HEADER_ALIASES["in_stock"])
    in_stock = _parse_bool(in_stock_raw)
    if in_stock is None:
        in_stock = bool(quantity and quantity > 0)

    currency_raw = _first_non_empty(normalized_row, _INVENTORY_HEADER_ALIASES["currency"])
    currency = str(currency_raw).strip().upper() if currency_raw is not None else "NGN"
    category_raw = _first_non_empty(normalized_row, _INVENTORY_HEADER_ALIASES["category"])
    category = str(category_raw).strip().lower() if category_raw is not None else "general"

    product: dict[str, object] = {
        "id": product_id,
        "name": name_value,
        "category": category or "general",
        "price": price,
        "currency": currency or "NGN",
        "in_stock": bool(in_stock),
        "data_tier": data_tier,
    }
    description_raw = _first_non_empty(normalized_row, _INVENTORY_HEADER_ALIASES["description"])
    if isinstance(description_raw, str) and description_raw.strip():
        product["description"] = description_raw.strip()
    brand_raw = _first_non_empty(normalized_row, _INVENTORY_HEADER_ALIASES["brand"])
    if isinstance(brand_raw, str) and brand_raw.strip():
        product["brand"] = brand_raw.strip()
    warehouse_raw = _first_non_empty(normalized_row, _INVENTORY_HEADER_ALIASES["warehouse"])
    if isinstance(warehouse_raw, str) and warehouse_raw.strip():
        product["warehouse"] = warehouse_raw.strip()
    if quantity is not None:
        product["quantity"] = quantity

    from app.configs.registry_schema import validate_product

    validation_errors = validate_product(product)
    if validation_errors:
        return None, f"row {index}: {'; '.join(validation_errors)}"
    return product, None


def _normalize_inventory_rows(
    rows: list[dict[str, object]],
    *,
    data_tier: str,
) -> tuple[list[dict[str, object]], list[str]]:
    products: list[dict[str, object]] = []
    errors: list[str] = []
    for index, row in enumerate(rows, start=1):
        normalized, error = _normalize_inventory_row(row, index=index, data_tier=data_tier)
        if error:
            errors.append(error)
            continue
        if normalized is not None:
            products.append(normalized)
    return products, errors


def _parse_csv_rows(raw: bytes) -> list[dict[str, object]]:
    text = raw.decode("utf-8-sig", errors="ignore")
    if not text.strip():
        return []
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    rows: list[dict[str, object]] = []
    for row in reader:
        if not isinstance(row, dict):
            continue
        normalized = {str(key): value for key, value in row.items() if key is not None}
        if not any(str(value or "").strip() for value in normalized.values()):
            continue
        rows.append(normalized)
    return rows


def _parse_xlsx_rows(raw: bytes, *, sheet_name: str | None) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("XLSX import requires openpyxl. Install openpyxl to enable this path.") from exc

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers_raw = next(iterator, None)
        if not headers_raw:
            return []
        headers = [str(value).strip() if value is not None else "" for value in headers_raw]
        rows: list[dict[str, object]] = []
        for values in iterator:
            if not values:
                continue
            row: dict[str, object] = {}
            has_data = False
            for idx, value in enumerate(values):
                header = headers[idx] if idx < len(headers) else ""
                if not header:
                    continue
                row[header] = value
                if value not in (None, ""):
                    has_data = True
            if has_data:
                rows.append(row)
        return rows
    finally:
        workbook.close()


def _parse_google_sheet_ref(source_url: str) -> tuple[str | None, str | None]:
    parsed = urlparse(source_url.strip())
    if parsed.scheme not in {"http", "https"}:
        return None, None
    if parsed.netloc not in {"docs.google.com", "www.docs.google.com"}:
        return None, None

    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", parsed.path)
    if match is None:
        return None, None
    spreadsheet_id = match.group(1)

    query_params = parse_qs(parsed.query, keep_blank_values=False)
    gid = query_params.get("gid", [None])[0]
    if not gid and parsed.fragment:
        fragment_gid = parse_qs(parsed.fragment, keep_blank_values=False).get("gid", [None])[0]
        gid = fragment_gid
    return spreadsheet_id, str(gid).strip() if isinstance(gid, str) and gid.strip() else None


def _google_sheet_csv_export_url(*, spreadsheet_id: str, gid: str | None) -> str:
    base = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv"
    if gid:
        return f"{base}&gid={gid}"
    return base


async def _sync_company_inventory_from_rows(
    *,
    tenant_id: str,
    company_id: str,
    rows: list[dict[str, object]],
    data_tier: str,
    dry_run: bool,
) -> dict[str, object]:
    products, normalization_errors = _normalize_inventory_rows(rows, data_tier=data_tier)
    operations: dict[str, int] = {
        "created": 0,
        "updated": 0,
        "unchanged": 0,
        "failed": len(normalization_errors),
    }

    if dry_run:
        operations["created"] = len(products)
        return {
            "parsedRows": len(rows),
            "normalizedRows": len(products),
            "written": 0,
            "operations": operations,
            "errors": normalization_errors,
            "dryRun": True,
        }

    import_result = await _import_company_products(
        tenant_id=tenant_id,
        company_id=company_id,
        products=products,
        data_tier=data_tier,
    )
    import_operations = import_result.get("operations")
    if isinstance(import_operations, dict):
        operations = {
            "created": int(import_operations.get("created", 0)),
            "updated": int(import_operations.get("updated", 0)),
            "unchanged": int(import_operations.get("unchanged", 0)),
            "failed": int(import_operations.get("failed", 0)) + len(normalization_errors),
        }
    import_errors = import_result.get("errors")
    if not isinstance(import_errors, list):
        import_errors = []
    return {
        "parsedRows": len(rows),
        "normalizedRows": len(products),
        "written": int(import_result.get("written", 0)),
        "operations": operations,
        "errors": normalization_errors + [str(error) for error in import_errors],
        "dryRun": False,
    }


async def _sync_company_inventory_from_google_sheet(
    *,
    tenant_id: str,
    company_id: str,
    source_url: str,
    sheet_name: str | None,
    data_tier: str,
    dry_run: bool,
) -> dict[str, object]:
    spreadsheet_id, gid = _parse_google_sheet_ref(source_url)
    if not spreadsheet_id:
        raise ValueError("Invalid Google Sheets URL")
    export_url = _google_sheet_csv_export_url(spreadsheet_id=spreadsheet_id, gid=gid)
    timeout_seconds = float(getattr(_m, "INVENTORY_SYNC_HTTP_TIMEOUT_SECONDS", 20.0))
    async with httpx.AsyncClient(timeout=timeout_seconds, follow_redirects=True) as client:
        response = await client.get(export_url)
    if response.status_code >= 400:
        raise RuntimeError(
            "Failed to read Google Sheet; ensure the sheet is shared or published. "
            f"status={response.status_code}"
        )
    raw = response.content
    max_bytes = int(getattr(_m, "INVENTORY_IMPORT_MAX_BYTES", 5_242_880))
    if len(raw) > max_bytes:
        raise ValueError(f"Inventory source exceeds max size ({max_bytes} bytes)")

    rows = _parse_csv_rows(raw)
    result = await _sync_company_inventory_from_rows(
        tenant_id=tenant_id,
        company_id=company_id,
        rows=rows,
        data_tier=data_tier,
        dry_run=dry_run,
    )
    result["sourceType"] = "google_sheets"
    result["sourceUrl"] = source_url
    result["sheetName"] = sheet_name or ""
    return result


async def _sync_company_inventory_from_connector(
    *,
    tenant_id: str,
    company_id: str,
    company_doc: dict[str, object],
    connector_id: str,
    data_tier: str,
    dry_run: bool,
) -> dict[str, object]:
    connectors = company_doc.get("connectors")
    connector_map = connectors if isinstance(connectors, dict) else {}
    connector = connector_map.get(connector_id)
    if not isinstance(connector, dict):
        raise ValueError("Connector not found")
    if connector.get("enabled") is False:
        raise ValueError("Connector is disabled")
    provider = str(connector.get("provider") or "").strip().lower()
    config = connector.get("config")
    config_dict = config if isinstance(config, dict) else {}

    if provider != "mock":
        raise NotImplementedError(
            "Connector inventory sync is currently available for provider='mock'. "
            "Configure mock connector config.inventory_rows to validate flow."
        )
    rows_raw = config_dict.get("inventory_rows")
    if not isinstance(rows_raw, list):
        raise ValueError("mock connector config.inventory_rows must be an array of objects")
    rows = [item for item in rows_raw if isinstance(item, dict)]
    result = await _sync_company_inventory_from_rows(
        tenant_id=tenant_id,
        company_id=company_id,
        rows=rows,
        data_tier=data_tier,
        dry_run=dry_run,
    )
    result["sourceType"] = "mcp_connector"
    result["connectorId"] = connector_id
    result["provider"] = provider
    return result


async def _sync_company_inventory_from_upload(
    *,
    tenant_id: str,
    company_id: str,
    filename: str,
    raw: bytes,
    sheet_name: str | None,
    data_tier: str,
    dry_run: bool,
) -> dict[str, object]:
    name = filename.strip().lower()
    if name.endswith(".xlsx"):
        rows = _parse_xlsx_rows(raw, sheet_name=sheet_name)
    else:
        rows = _parse_csv_rows(raw)
    result = await _sync_company_inventory_from_rows(
        tenant_id=tenant_id,
        company_id=company_id,
        rows=rows,
        data_tier=data_tier,
        dry_run=dry_run,
    )
    result["sourceType"] = "file_upload"
    result["fileName"] = filename
    if sheet_name:
        result["sheetName"] = sheet_name
    return result


def _inventory_sync_metadata(
    *,
    source_type: str,
    data_tier: str,
    dry_run: bool,
    source_url: str | None = None,
    connector_id: str | None = None,
    sheet_name: str | None = None,
    status: str,
    written: int,
    parsed_rows: int,
    normalized_rows: int,
    errors: list[str],
    auto_enabled: bool | None = None,
    interval_minutes: int | None = None,
    next_run_at: str | None = None,
    last_attempt_at: str | None = None,
    last_error: str | None = None,
) -> dict[str, object]:
    metadata: dict[str, object] = {
        "source_type": source_type,
        "data_tier": data_tier,
        "dry_run": dry_run,
        "status": status,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "last_result": {
            "written": written,
            "parsed_rows": parsed_rows,
            "normalized_rows": normalized_rows,
            "error_count": len(errors),
        },
    }
    if source_url:
        metadata["source_url"] = source_url
    if connector_id:
        metadata["connector_id"] = connector_id
    if sheet_name:
        metadata["sheet_name"] = sheet_name
    if auto_enabled is not None:
        metadata["auto_enabled"] = bool(auto_enabled)
    if interval_minutes is not None:
        metadata["interval_minutes"] = max(1, min(int(interval_minutes), 1440))
    if isinstance(next_run_at, str) and next_run_at.strip():
        metadata["next_run_at"] = next_run_at.strip()
    if isinstance(last_attempt_at, str) and last_attempt_at.strip():
        metadata["last_attempt_at"] = last_attempt_at.strip()
    if isinstance(last_error, str) and last_error.strip():
        metadata["last_error"] = last_error.strip()
    return metadata


def _inventory_sync_interval_minutes(metadata: dict[str, object]) -> int:
    try:
        value = int(metadata.get("interval_minutes", 15))
    except (TypeError, ValueError):
        value = 15
    return max(1, min(value, 1440))


def _inventory_sync_due(
    metadata: dict[str, object],
    *,
    now_utc: datetime,
    force: bool,
) -> bool:
    if force:
        return True
    if not bool(metadata.get("auto_enabled", False)):
        return False
    next_run_at_raw = metadata.get("next_run_at")
    next_run_at = _parse_timestamp_utc(next_run_at_raw)
    if next_run_at is None:
        return True
    return next_run_at <= now_utc


def _inventory_sync_next_run_at(*, now_utc: datetime, interval_minutes: int) -> str:
    interval = max(1, min(interval_minutes, 1440))
    return (now_utc + timedelta(minutes=interval)).isoformat()


async def _list_tenant_company_docs(
    *,
    tenant_id: str,
    max_companies: int,
) -> list[tuple[str, dict[str, object]]]:
    db = _m._registry_db_client()
    if db is None:
        raise RuntimeError("Registry storage unavailable")
    collection_ref = (
        db.collection("tenants")
        .document(tenant_id)
        .collection("companies")
        .limit(max(1, min(max_companies, 500)))
    )
    docs = await _collect_query_docs(collection_ref)
    companies: list[tuple[str, dict[str, object]]] = []
    for doc in docs:
        company_id = str(getattr(doc, "id", "")).strip().lower()
        if not company_id:
            continue
        payload = doc.to_dict() if hasattr(doc, "to_dict") else {}
        if not isinstance(payload, dict):
            payload = {}
        companies.append((company_id, payload))
    return companies


async def _run_inventory_sync_jobs(
    *,
    tenant_id: str,
    company_id: str | None = None,
    max_companies: int = 50,
    force: bool = False,
    dry_run_override: bool | None = None,
) -> dict[str, object]:
    now_utc = datetime.now(timezone.utc)
    now_iso = now_utc.isoformat()
    target_company_id = (company_id or "").strip().lower() or None

    if target_company_id:
        company_doc, company_error = await _m._load_registry_company_doc(
            tenant_id=tenant_id,
            company_id=target_company_id,
        )
        if company_error:
            raise ValueError(f"Company not found for tenant: {target_company_id}")
        companies = [(target_company_id, company_doc if isinstance(company_doc, dict) else {})]
    else:
        companies = await _list_tenant_company_docs(
            tenant_id=tenant_id,
            max_companies=max_companies,
        )

    report: list[dict[str, object]] = []
    triggered = 0
    skipped = 0

    for normalized_company_id, company_doc in companies:
        inventory_sync = company_doc.get("inventory_sync")
        metadata = inventory_sync if isinstance(inventory_sync, dict) else {}
        source_type = str(metadata.get("source_type") or "").strip().lower()
        if source_type not in {"google_sheets", "mcp_connector"}:
            skipped += 1
            report.append(
                {
                    "companyId": normalized_company_id,
                    "status": "skipped",
                    "reason": "No supported inventory sync source configured",
                }
            )
            continue
        if not _inventory_sync_due(metadata, now_utc=now_utc, force=force):
            skipped += 1
            report.append(
                {
                    "companyId": normalized_company_id,
                    "status": "skipped",
                    "reason": "Not due",
                }
            )
            continue

        data_tier = str(metadata.get("data_tier") or "admin").strip().lower() or "admin"
        effective_dry_run = bool(dry_run_override) if dry_run_override is not None else bool(
            metadata.get("dry_run", False)
        )
        source_url = str(metadata.get("source_url") or "").strip()
        connector_id = str(metadata.get("connector_id") or "").strip().lower()
        sheet_name = str(metadata.get("sheet_name") or "").strip() or None

        status = "success"
        result: dict[str, object] = {
            "parsedRows": 0,
            "normalizedRows": 0,
            "written": 0,
            "errors": [],
        }
        triggered += 1
        try:
            if source_type == "google_sheets":
                if not source_url:
                    raise ValueError("Missing source_url for google_sheets")
                result = await _sync_company_inventory_from_google_sheet(
                    tenant_id=tenant_id,
                    company_id=normalized_company_id,
                    source_url=source_url,
                    sheet_name=sheet_name,
                    data_tier=data_tier,
                    dry_run=effective_dry_run,
                )
            else:
                if not connector_id:
                    raise ValueError("Missing connector_id for mcp_connector")
                result = await _sync_company_inventory_from_connector(
                    tenant_id=tenant_id,
                    company_id=normalized_company_id,
                    company_doc=company_doc,
                    connector_id=connector_id,
                    data_tier=data_tier,
                    dry_run=effective_dry_run,
                )
            result_errors = result.get("errors")
            if isinstance(result_errors, list) and result_errors:
                status = "partial"
        except Exception as exc:
            status = "error"
            result = {
                "parsedRows": 0,
                "normalizedRows": 0,
                "written": 0,
                "errors": [str(exc)],
            }

        interval_minutes = _inventory_sync_interval_minutes(metadata)
        auto_enabled = bool(metadata.get("auto_enabled", False))
        next_run_at = (
            _inventory_sync_next_run_at(now_utc=now_utc, interval_minutes=interval_minutes)
            if auto_enabled
            else None
        )
        result_errors = result.get("errors")
        normalized_errors = (
            [str(item) for item in result_errors]
            if isinstance(result_errors, list)
            else []
        )
        merged_metadata = _inventory_sync_metadata(
            source_type=source_type,
            source_url=source_url or None,
            connector_id=connector_id or None,
            sheet_name=sheet_name,
            data_tier=data_tier,
            dry_run=effective_dry_run,
            status=status,
            written=int(result.get("written", 0)),
            parsed_rows=int(result.get("parsedRows", 0)),
            normalized_rows=int(result.get("normalizedRows", 0)),
            errors=normalized_errors,
            auto_enabled=auto_enabled,
            interval_minutes=interval_minutes,
            next_run_at=next_run_at,
            last_attempt_at=now_iso,
            last_error=normalized_errors[0] if status == "error" and normalized_errors else None,
        )
        configured_at = metadata.get("configured_at")
        if isinstance(configured_at, str) and configured_at.strip():
            merged_metadata["configured_at"] = configured_at.strip()

        await _m._save_registry_company_doc(
            tenant_id=tenant_id,
            company_id=normalized_company_id,
            payload={
                "inventory_sync": merged_metadata,
                "updated_at": now_iso,
            },
        )
        report.append(
            {
                "companyId": normalized_company_id,
                "status": status,
                "written": int(result.get("written", 0)),
                "parsedRows": int(result.get("parsedRows", 0)),
                "normalizedRows": int(result.get("normalizedRows", 0)),
                "errorCount": len(normalized_errors),
            }
        )

    return {
        "tenantId": tenant_id,
        "companyId": target_company_id,
        "force": bool(force),
        "dryRunOverride": dry_run_override,
        "processed": len(companies),
        "triggered": triggered,
        "skipped": skipped,
        "results": report,
        "runAt": now_iso,
    }


async def _import_company_runtime_docs(
    *,
    tenant_id: str,
    company_id: str,
    collection_name: str,
    items: list[dict[str, object]],
    data_tier: str,
    validator: object,
) -> dict[str, object]:
    db = _m._registry_db_client()
    if db is None:
        raise RuntimeError("Registry storage unavailable")

    operations = {"created": 0, "updated": 0, "unchanged": 0, "failed": 0}
    errors: list[str] = []
    doc_payloads: list[tuple[object, dict[str, object]]] = []
    validate = validator if callable(validator) else None

    for item in items:
        if not isinstance(item, dict):
            operations["failed"] += 1
            errors.append("entry: item must be an object")
            continue
        normalized = dict(item)
        entry_id = str(normalized.get("id", "")).strip()
        if not entry_id:
            operations["failed"] += 1
            errors.append("entry: missing required field 'id'")
            continue
        if "data_tier" not in normalized:
            normalized["data_tier"] = data_tier
        validation_errors = validate(normalized) if validate else []
        if validation_errors:
            operations["failed"] += 1
            errors.append(f"entry '{entry_id}': {'; '.join(validation_errors)}")
            continue

        doc_ref = (
            db.collection("tenants")
            .document(tenant_id)
            .collection("companies")
            .document(company_id)
            .collection(collection_name)
            .document(entry_id)
        )
        existing_snapshot = await _doc_get(doc_ref)
        existing_data = (
            existing_snapshot.to_dict()
            if getattr(existing_snapshot, "exists", False) and hasattr(existing_snapshot, "to_dict")
            else {}
        )
        if existing_data == normalized:
            operations["unchanged"] += 1
            continue
        if getattr(existing_snapshot, "exists", False):
            operations["updated"] += 1
        else:
            operations["created"] += 1
        doc_payloads.append((doc_ref, normalized))

    for start in range(0, len(doc_payloads), 500):
        chunk = doc_payloads[start : start + 500]
        await _batch_set_documents(db, chunk, merge=False)

    return {"written": len(doc_payloads), "operations": operations, "errors": errors}


async def _import_company_products(
    *,
    tenant_id: str,
    company_id: str,
    products: list[dict[str, object]],
    data_tier: str = "admin",
) -> dict[str, object]:
    from app.configs.registry_schema import validate_product

    return await _import_company_runtime_docs(
        tenant_id=tenant_id,
        company_id=company_id,
        collection_name="products",
        items=products,
        data_tier=data_tier,
        validator=validate_product,
    )


async def _import_company_booking_slots(
    *,
    tenant_id: str,
    company_id: str,
    slots: list[dict[str, object]],
    data_tier: str = "admin",
) -> dict[str, object]:
    from app.configs.registry_schema import validate_booking_slot

    return await _import_company_runtime_docs(
        tenant_id=tenant_id,
        company_id=company_id,
        collection_name="booking_slots",
        items=slots,
        data_tier=data_tier,
        validator=validate_booking_slot,
    )


async def _upsert_company_demo_connector(
    *,
    tenant_id: str,
    company_id: str,
    company_doc: dict[str, object],
) -> dict[str, object]:
    from app.api.models import AdminConnectorPayload

    connector_id = _m._normalize_connector_id(_DEMO_SEED_CONNECTOR.get("connectorId"))
    if not connector_id:
        raise ValueError("Demo seed connectorId is invalid")

    payload = AdminConnectorPayload(
        connectorId=connector_id,
        provider=str(_DEMO_SEED_CONNECTOR.get("provider") or "mock"),
        enabled=bool(_DEMO_SEED_CONNECTOR.get("enabled", True)),
        capabilities=list(_DEMO_SEED_CONNECTOR.get("capabilities") or []),
        config=dict(_DEMO_SEED_CONNECTOR.get("config") or {}),
    )
    connector, connector_error = _m._normalize_connector_payload(
        connector_id=connector_id,
        payload=payload,
        industry_template_id=str(company_doc.get("industry_template_id") or ""),
    )
    if connector_error:
        raise ValueError(
            _json_response_error_message(
                connector_error,
                fallback="Demo seed connector rejected by template/provider policy",
            )
        )
    if not isinstance(connector, dict):
        raise RuntimeError("Demo seed connector normalization failed")

    connectors_raw = company_doc.get("connectors")
    connectors = dict(connectors_raw) if isinstance(connectors_raw, dict) else {}
    existing = connectors.get(connector_id)
    if isinstance(existing, dict) and existing == connector:
        return {
            "connectorId": connector_id,
            "created": False,
            "updated": False,
            "unchanged": True,
        }

    connectors[connector_id] = connector
    await _m._save_registry_company_doc(
        tenant_id=tenant_id,
        company_id=company_id,
        payload={
            "connectors": connectors,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        },
    )

    return {
        "connectorId": connector_id,
        "created": not isinstance(existing, dict),
        "updated": isinstance(existing, dict),
        "unchanged": False,
    }


async def _seed_company_demo_data(
    *,
    tenant_id: str,
    company_id: str,
    company_doc: dict[str, object],
    data_tier: str = "demo",
) -> dict[str, object]:
    from app.configs.registry_schema import validate_knowledge_entry

    safe_data_tier = str(data_tier or "demo").strip().lower() or "demo"
    knowledge_entries = _demo_seed_knowledge_entries(data_tier=safe_data_tier)
    products = [{**item, "data_tier": safe_data_tier} for item in _DEMO_SEED_PRODUCTS]
    booking_slots = [{**item, "data_tier": safe_data_tier} for item in _DEMO_SEED_BOOKING_SLOTS]

    section_tasks = {
        "knowledge": _import_company_runtime_docs(
            tenant_id=tenant_id,
            company_id=company_id,
            collection_name="knowledge",
            items=knowledge_entries,
            data_tier=safe_data_tier,
            validator=validate_knowledge_entry,
        ),
        "connectors": _upsert_company_demo_connector(
            tenant_id=tenant_id,
            company_id=company_id,
            company_doc=company_doc,
        ),
        "products": _import_company_products(
            tenant_id=tenant_id,
            company_id=company_id,
            products=products,
            data_tier=safe_data_tier,
        ),
        "booking_slots": _import_company_booking_slots(
            tenant_id=tenant_id,
            company_id=company_id,
            slots=booking_slots,
            data_tier=safe_data_tier,
        ),
    }

    results = await asyncio.gather(*section_tasks.values(), return_exceptions=True)
    sections: dict[str, dict[str, object]] = {}
    errors: list[str] = []

    for section_name, result in zip(section_tasks.keys(), results):
        if isinstance(result, Exception):
            message = str(result).strip() or "Seed failed"
            sections[section_name] = {"ok": False, "error": message}
            errors.append(f"{section_name}: {message}")
            continue
        payload = dict(result) if isinstance(result, dict) else {"value": result}
        payload["ok"] = True
        sections[section_name] = payload

    ok = len(errors) == 0
    return {
        "seedVersion": _DEMO_SEED_VERSION,
        "dataTier": safe_data_tier,
        "ok": ok,
        "sections": sections,
        "errors": errors,
    }


async def _purge_company_demo_runtime_data(
    *,
    tenant_id: str,
    company_id: str,
) -> dict[str, int]:
    db = _m._registry_db_client()
    if db is None:
        raise RuntimeError("Registry storage unavailable")

    deleted = {"products": 0, "booking_slots": 0, "knowledge": 0}
    for subcollection in ("products", "booking_slots", "knowledge"):
        query = (
            db.collection("tenants")
            .document(tenant_id)
            .collection("companies")
            .document(company_id)
            .collection(subcollection)
            .where("data_tier", "==", "demo")
        )
        docs = await _collect_query_docs(query)
        doc_refs: list[object] = []
        for doc in docs:
            doc_ref = getattr(doc, "reference", None)
            if doc_ref is None:
                continue
            doc_refs.append(doc_ref)
        for start in range(0, len(doc_refs), 500):
            chunk = doc_refs[start : start + 500]
            await _batch_delete_documents(db, chunk)
            deleted[subcollection] += len(chunk)
    return deleted


def _parse_timestamp_utc(raw_value: object) -> datetime | None:
    if not isinstance(raw_value, str):
        return None
    value = raw_value.strip()
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


async def _export_company_bundle(
    *,
    tenant_id: str,
    company_id: str,
    company_doc: dict[str, object],
    include_runtime_data: bool,
) -> dict[str, object]:
    knowledge_entries = await _list_company_collection_docs(
        tenant_id=tenant_id,
        company_id=company_id,
        collection_name="knowledge",
    )
    products: list[dict[str, object]] = []
    booking_slots: list[dict[str, object]] = []
    if include_runtime_data:
        products = await _list_company_collection_docs(
            tenant_id=tenant_id,
            company_id=company_id,
            collection_name="products",
        )
        booking_slots = await _list_company_collection_docs(
            tenant_id=tenant_id,
            company_id=company_id,
            collection_name="booking_slots",
        )
    return {
        "company": _admin_company_response(
            tenant_id=tenant_id,
            company_id=company_id,
            raw_company=company_doc,
        ),
        "collections": {
            "knowledge": knowledge_entries,
            "products": products,
            "booking_slots": booking_slots,
        },
        "counts": {
            "knowledge": len(knowledge_entries),
            "products": len(products),
            "booking_slots": len(booking_slots),
        },
    }


async def _delete_company_bundle(
    *,
    tenant_id: str,
    company_id: str,
) -> dict[str, int]:
    db = _m._registry_db_client()
    if db is None:
        raise RuntimeError("Registry storage unavailable")

    deleted_counts = {"knowledge": 0, "products": 0, "booking_slots": 0, "company": 0}
    company_doc_ref = (
        db.collection("tenants")
        .document(tenant_id)
        .collection("companies")
        .document(company_id)
    )
    for subcollection in ("knowledge", "products", "booking_slots"):
        collection_ref = company_doc_ref.collection(subcollection)
        docs = await _collect_query_docs(collection_ref)
        doc_refs: list[object] = []
        for doc in docs:
            doc_ref = getattr(doc, "reference", None)
            if doc_ref is None:
                continue
            doc_refs.append(doc_ref)
        for start in range(0, len(doc_refs), 500):
            chunk = doc_refs[start : start + 500]
            await _batch_delete_documents(db, chunk)
            deleted_counts[subcollection] += len(chunk)

    snapshot = await _doc_get(company_doc_ref)
    if getattr(snapshot, "exists", False):
        await _doc_delete(company_doc_ref)
        deleted_counts["company"] = 1
    return deleted_counts


async def _purge_company_retention_data(
    *,
    tenant_id: str,
    company_id: str,
    older_than_days: int,
    collections: list[str],
    data_tier: str | None = None,
) -> dict[str, dict[str, int]]:
    db = _m._registry_db_client()
    if db is None:
        raise RuntimeError("Registry storage unavailable")

    cutoff = datetime.now(timezone.utc) - timedelta(days=max(1, older_than_days))
    target_tier = (data_tier or "").strip().lower() or None
    report: dict[str, dict[str, int]] = {}
    company_doc_ref = (
        db.collection("tenants")
        .document(tenant_id)
        .collection("companies")
        .document(company_id)
    )

    for collection_name in collections:
        counters = {"scanned": 0, "deleted": 0, "skipped": 0, "missing_timestamp": 0}
        collection_ref = company_doc_ref.collection(collection_name)
        docs = await _collect_query_docs(collection_ref)
        deletable_refs: list[object] = []
        for doc in docs:
            counters["scanned"] += 1
            item = doc.to_dict() if hasattr(doc, "to_dict") else {}
            if not isinstance(item, dict):
                item = {}
            if target_tier:
                item_tier = str(item.get("data_tier", "")).strip().lower()
                if item_tier != target_tier:
                    counters["skipped"] += 1
                    continue
            timestamp = _parse_timestamp_utc(item.get("updated_at")) or _parse_timestamp_utc(
                item.get("created_at")
            )
            if timestamp is None:
                counters["missing_timestamp"] += 1
                counters["skipped"] += 1
                continue
            if timestamp >= cutoff:
                counters["skipped"] += 1
                continue
            doc_ref = getattr(doc, "reference", None)
            if doc_ref is None:
                counters["skipped"] += 1
                continue
            deletable_refs.append(doc_ref)
        for start in range(0, len(deletable_refs), 500):
            chunk = deletable_refs[start : start + 500]
            await _batch_delete_documents(db, chunk)
            counters["deleted"] += len(chunk)
        report[collection_name] = counters
    return report
